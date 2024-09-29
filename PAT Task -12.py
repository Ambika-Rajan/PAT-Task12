import openpyxl
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pytest
from selenium import webdriver
from login_page import LoginPage


# Create an Excel workbook and add a sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Login Test Data"

# Define the headers
headers = ["Test ID", "Username", "Password", "Date", "Time of Test", "Name of Tester", "Test Result"]
sheet.append(headers)

# Sample data for 5 test cases
test_data = [
    [1, "Admin", "admin123", "", "", "Tester A", ""],
    [2, "user1", "password1", "", "", "Tester A", ""],
    [3, "user2", "password2", "", "", "Tester A", ""],
    [4, "user3", "password3", "", "", "Tester A", ""],
    [5, "user4", "password4", "", "", "Tester A", ""],
]

for row in test_data:
    sheet.append(row)

# Save the Excel file
workbook.save("test_data.xlsx")

class LoginPage:
    def __init__(self, driver):
        self.driver = driver
        self.username_input = (By.NAME, "username")
        self.password_input = (By.NAME, "password")
        self.login_button = (By.XPATH, "//button[@type='submit']")
        self.logout_button = (By.XPATH, "//a[text()='Logout']")

    def login(self, username, password):
        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(self.username_input)).send_keys(username)
        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(self.password_input)).send_keys(password)
        WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(self.login_button)).click()

    def is_logged_in(self):
        try:
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(self.logout_button))
            return True
        except:
            return False

        # Load test data from Excel
        def load_test_data(file_name):
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            test_cases = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                test_cases.append(list(row))
            return test_cases

        @pytest.mark.parametrize("test_data", load_test_data("test_data.xlsx"))
        def test_login(test_data):
            test_id, username, password, date, time, tester_name, result = test_data

            # Initialize WebDriver
            driver = webdriver.Chrome()  # Ensure you have the Chrome WebDriver installed
            driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")

            # Create an instance of the LoginPage
            login_page = LoginPage(driver)

            # Perform login
            login_page.login(username, password)

            # Check if login was successful
            if login_page.is_logged_in():
                result = "Passed"
            else:
                result = "Failed"

            # Update the Excel file with the result
            date = datetime.now().date()
            time = datetime.now().time().strftime("%H:%M:%S")
            update_test_result("test_data.xlsx", test_id, date, time, tester_name, result)

            # Close the browser
            driver.quit()

        def update_test_result(file_name, test_id, date, time, tester_name, result):
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2):
                if row[0].value == test_id:
                    row[3].value = date
                    row[4].value = time
                    row[5].value = tester_name
                    row[6].value = result
                    break

            workbook.save(file_name)